import self as self
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl as xl
from reportlab.platypus import BaseDocTemplate
from reportlab.rl_settings import defaultPageSize

from plot_data import *
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from PIL import Image

import test

# loading excel file
wb = xl.load_workbook('Test.xlsx')
sheet = wb['Number']

max_row = sheet.max_row

# finding maximum column number
max_col = 1
while sheet.cell(1, max_col).value is not None:
    max_col += 1
else:
    max_col -= 1

# creating an empty list
data_list = [None] * max_row
for i in range(max_row):
    data_list[i] = [None] * max_col

# transfer data from sheet to the 2-D list
for row_count in range(max_row):
    for col_count in range(max_col):
        data_list[row_count][col_count] = sheet.cell(row_count + 1, col_count + 1).value

for i in range(max_row):
    print(data_list[i][:])

# create a PdfPages object
c = canvas.Canvas("Out2.pdf")
c.drawString(100, 100, "Hello World")

# plot all parameters
for col_number in range(5, 6, 3):
    fig = plot_data_col(data_list, col_number)
    # save the current figure
    fig.savefig("temp.png")
    test.build_pdf("Out3.pdf")
    img = Image.open("temp.png")
    print(img.width, img.height)
    c.drawImage("temp.png", 0 * mm, 0 * mm, 100 * mm, 100 * img.height / img.width * mm)

c.showPage()
c.save()
