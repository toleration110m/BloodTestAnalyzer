"""
excel_reader.py

Read blood test data from an Excel workbook.
"""

from pathlib import Path

import openpyxl

from config import INPUT_FILE, SHEET_NAME
from models import ParameterData


def read_excel(file_path: Path = INPUT_FILE) -> list[ParameterData]:
    """
    Read the Excel workbook and return all blood test parameters.

    Expected worksheet format:

    ------------------------------------------------------------------
    | Date | ... | Low | High | Parameter | Low | High | Parameter ...
    ------------------------------------------------------------------

    Every parameter occupies three consecutive columns:
        Lower Limit | Upper Limit | Measured Value
    """

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[SHEET_NAME]

    max_row = sheet.max_row
    max_col = sheet.max_column

    parameters: list[ParameterData] = []

    #
    # Parameter columns start at column 6 (Excel column F)
    #
    # Layout:
    #
    # A : Date
    # ...
    # D : Lower
    # E : Upper
    # F : Parameter
    # G : Lower
    # H : Upper
    # I : Parameter
    #
    # Therefore:
    #
    # parameter columns = 6, 9, 12, ...
    #

    for parameter_col in range(6, max_col + 1, 3):

        parameter_name = sheet.cell(row=1, column=parameter_col).value

        if parameter_name is None:
            continue

        parameter = ParameterData(name=str(parameter_name))

        lower_col = parameter_col - 2
        upper_col = parameter_col - 1

        for row in range(2, max_row + 1):

            date = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=parameter_col).value
            lower = sheet.cell(row=row, column=lower_col).value
            upper = sheet.cell(row=row, column=upper_col).value

            if value is None:
                continue

            parameter.dates.append(date)
            parameter.values.append(value)
            parameter.lower_limits.append(lower)
            parameter.upper_limits.append(upper)

        parameters.append(parameter)

    workbook.close()

    return parameters